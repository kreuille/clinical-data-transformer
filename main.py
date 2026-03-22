import os, re, json, tempfile, base64
import functions_framework
from flask import jsonify, request
import google.generativeai as genai
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY')
GEMINI_MODEL = os.environ.get('GEMINI_MODEL', 'gemini-2.5-flash')
ALLOWED_ORIGINS = os.environ.get('ALLOWED_ORIGINS', 'https://kreuille.github.io')
genai.configure(api_key=GOOGLE_API_KEY)

SYSTEM_PROMPT = """You are an expert Python/openpyxl developer for Excel transformations.
You receive: column names, 3 sample rows, and a user instruction.
RESPOND WITH ONLY VALID JSON. No markdown, no code fences.
The JSON has 3 fields:

1. "server_code": Python code to transform an Excel file ON THE SERVER.
   Variables already available: input_path, output_path (file paths as strings).
   Libraries already imported: pd, openpyxl, PatternFill, Font, Alignment, Border, Side, BarChart, LineChart, PieChart, Reference, dataframe_to_rows.
   RULES: Do NOT import anything. Do NOT use tkinter. Do NOT use print().
   Read input: df = pd.read_excel(input_path, engine='openpyxl')
   Apply ALL transformations: data changes, colors, formatting, charts, conditional formatting.
   Save to output_path using openpyxl Workbook.
   ALWAYS save: wb.save(output_path)

2. "python_script": Complete standalone Python script for LOCAL execution.
   Import everything needed. Use tkinter.filedialog for file selection.
   Apply the SAME transformations. French comments. Error handling.

3. "js_code": Simple JS function body for HTML preview.
   Receives 'data' array, returns transformed array. var/function() only.
   For visual-only changes (colors/charts): just return data;

For MODIFICATIONS: combine ALL previous + new into standalone code.
ALWAYS return valid JSON only."""

def _cors():
    origin = request.headers.get('Origin', '')
    allowed = [o.strip() for o in ALLOWED_ORIGINS.split(',')]
    h = {'Access-Control-Allow-Methods': 'POST, OPTIONS', 'Access-Control-Allow-Headers': 'Content-Type', 'Access-Control-Max-Age': '3600'}
    h['Access-Control-Allow-Origin'] = origin if origin in allowed else ''
    return h

@functions_framework.http
def clinical_transform(req):
    cors = _cors()
    if req.method == 'OPTIONS':
        return ('', 204, cors)
    try:
        data = req.get_json(silent=True)
        if not data:
            return (jsonify(error='JSON invalide.'), 400, cors)
        columns = data.get('columns')
        sample = data.get('sample')
        all_data = data.get('all_data')
        instruction = data.get('instruction', '')
        history = data.get('history', [])
        if not columns or not sample:
            return (jsonify(error='Champs requis manquants.'), 400, cors)

        parts = [
            'Colonnes: ' + json.dumps(columns, ensure_ascii=False),
            'Echantillon (3 lignes): ' + json.dumps(sample, ensure_ascii=False)
        ]
        if history:
            parts.append('=== HISTORIQUE ===')
            for item in history:
                parts.append(item['role'].upper() + ': ' + item['content'])
            parts.append('=== FIN HISTORIQUE ===')
        parts.append('INSTRUCTION: ' + instruction)
        user_msg = '\n'.join(parts)

        model = genai.GenerativeModel(model_name=GEMINI_MODEL, system_instruction=SYSTEM_PROMPT)
        response = model.generate_content(user_msg)
        txt = response.text.strip()
        txt = re.sub(r'^```(?:json)?\n?', '', txt)
        txt = re.sub(r'\n?```$', '', txt)
        try:
            result = json.loads(txt)
        except json.JSONDecodeError:
            return (jsonify(error='Reponse IA invalide', raw=txt), 500, cors)

        server_code = result.get('server_code', '')
        python_script = result.get('python_script', '')
        js_code = result.get('js_code', 'return data;')

        excel_b64 = ''
        exec_error = ''
        if all_data and server_code:
            tmp_in_path = ''
            tmp_out_path = ''
            try:
                df_input = pd.DataFrame(all_data)
                tmp_in = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                tmp_in_path = tmp_in.name
                tmp_in.close()
                df_input.to_excel(tmp_in_path, index=False, engine='openpyxl')
                tmp_out = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                tmp_out_path = tmp_out.name
                tmp_out.close()
                namespace = {
                    'pd': pd, 'openpyxl': openpyxl,
                    'PatternFill': PatternFill, 'Font': Font, 'Alignment': Alignment,
                    'Border': Border, 'Side': Side,
                    'BarChart': BarChart, 'LineChart': LineChart, 'PieChart': PieChart,
                    'Reference': Reference, 'dataframe_to_rows': dataframe_to_rows,
                    'input_path': tmp_in_path, 'output_path': tmp_out_path,
                }
                exec(server_code, namespace)
                with open(tmp_out_path, 'rb') as f:
                    excel_b64 = base64.b64encode(f.read()).decode()
            except Exception as e:
                exec_error = str(e)
            finally:
                for p in [tmp_in_path, tmp_out_path]:
                    if p and os.path.exists(p):
                        os.unlink(p)

        resp = {'js_code': js_code, 'python_script': python_script, 'excel_b64': excel_b64}
        if exec_error:
            resp['exec_error'] = exec_error
        return (jsonify(**resp), 200, cors)

    except Exception as exc:
        return (jsonify(error=str(exc)), 500, cors)